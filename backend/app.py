from pyngrok import ngrok
from flask import Flask, request, jsonify
import json
import os
import time
from datetime import datetime
import pandas as pd
import requests
import pickle
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from google.auth.transport.requests import Request
import re
import pdfplumber
import traceback
from google_auth_oauthlib.flow import InstalledAppFlow
from openpyxl.utils import get_column_letter
from google.auth import exceptions

log_file_path = "/content/drive/My Drive/resume_parser_log.txt"

SPREADSHEET_ID = os.getenv("SPREADSHEET_ID", "1Yroi1LXhj4XvvVwCo1Lp349Ppi_21icswmP42uuA2nA")
SHEET_NAME = os.getenv("SHEET_NAME", "employees") 
SCOPES = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive.readonly']
API_NAME = 'sheets'
API_VERSION = 'v4'
SERVICE_ACCOUNT_FILE = '/content/credentials.json'

ngrok.set_auth_token("2s5CTZQqN4J5bp5pWnrgyZp2sNy_3K5To9pcnRnegTCPcwGr6")
app = Flask(__name__)

def authenticate_google_sheets():
    """Authenticate using a service account and return a Sheets API service."""
    try:
        creds = Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        service = build('sheets', 'v4', credentials=creds)
        print("Google Sheets API authenticated successfully!")
        return service
    except Exception as e:
        print("Google Sheets Authentication Failed:", str(e))
        return None

def download_pdf_from_drive(drive_url):
    """Download a PDF file from Google Drive given its URL."""
    try:
        if 'drive.google.com' not in drive_url:
            return None

        file_id = None
        if '/d/' in drive_url:
            file_id = drive_url.split('/d/')[1].split('/')[0]
        elif 'id=' in drive_url:
            file_id = drive_url.split('id=')[1].split('&')[0]

        if not file_id:
            return None

        download_url = f'https://drive.google.com/uc?export=download&id={file_id}'
        response = requests.get(download_url, allow_redirects=True)

        if response.status_code == 200 and b'%PDF' in response.content[:4]:
            pdf_path = 'temp_resume.pdf'
            with open(pdf_path, 'wb') as f:
                f.write(response.content)
            return pdf_path
        return None
    except Exception:
        return None

def extract_text_from_pdf(pdf_path):
    """Extract text from a PDF file."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            full_text = "".join([page.extract_text() for page in pdf.pages])
            return full_text
    except Exception:
        return None

def load_employee_data(service):
    """Load employee data from Google Sheets."""
    range_ = f"{SHEET_NAME}!A1:Z1000"
    try:
        result = service.spreadsheets().values().get(spreadsheetId=SPREADSHEET_ID, range=range_).execute()
        values = result.get('values', [])
        if not values:
            return pd.DataFrame()
        headers = values[0]
        data = values[1:]
        df = pd.DataFrame(data, columns=headers)
        return df
    except Exception:
        return pd.DataFrame()

def clean_and_truncate_data(df, max_length=500):
    """Clean and truncate DataFrame values to fit within Google Sheets limits."""
    for col in df.columns:
        df[col] = df[col].astype(str).apply(lambda x: (x[:max_length] + "...") if len(x) > max_length else x)
    return df

def get_range_for_columns(start_col, num_columns, num_rows):
    end_col = get_column_letter(start_col + num_columns - 1)
    return f"{get_column_letter(start_col)}1:{end_col}{num_rows + 1}"

def convert_to_string(value):
    if isinstance(value, dict) or isinstance(value, list):
        return str(value)
    return value

def update_google_sheet_with_parsed_data(service, spreadsheet_id, sheet_name, parsed_df):
    if parsed_df.empty:
        return
    required_columns = ["Experience", "Skills", "Department", "Projects", "Email", "LinkedIn", "CGPA", "Marks", "Achievements", "Other Info"]
    for column in required_columns:
        if column not in parsed_df.columns:
            parsed_df[column] = ""

    parsed_df = parsed_df[required_columns] 
    parsed_df = parsed_df.map(convert_to_string)
    
    for column in required_columns:
        parsed_df[column] = parsed_df[column].str.replace('\n', ' ', regex=True)

    parsed_df.fillna('', inplace=True)

    sheet_data = [parsed_df.columns.tolist()] + parsed_df.values.tolist()

    start_col = 9
    num_columns = len(parsed_df.columns)
    num_rows = len(parsed_df)
    range_ = get_range_for_columns(start_col, num_columns, num_rows)
    body = {'values': sheet_data}

    try:
        service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=range_,
            valueInputOption="RAW",
            body=body
        ).execute()
    except Exception as e:
        write_log(f"Error updating Google Sheet: {str(e)}")

def extract_section(text, keywords):
    """Extract sections from resume text based on keywords."""
    extracted_data = {}
    for keyword in keywords:
        keyword_lower = keyword.lower()
        if isinstance(text, str) and keyword_lower in text.lower():
            try:
                section_start = text.lower().index(keyword_lower)
                section_end = text.find("\n", section_start + len(keyword))
                if section_end == -1 or section_end - section_start < 50:
                    section_end = min(len(text), section_start + 500)
                extracted_data[keyword] = text[section_start:section_end].strip()
            except Exception:
                extracted_data[keyword] = "Not found"
        else:
            extracted_data[keyword] = "Not found"
    return extracted_data

def ensure_columns_exist(df):
    """Ensure that necessary columns exist."""
    columns_to_add = [
        'Experience', 'Skills', 'Projects', 'Department', 
        'Email', 'LinkedIn', 'CGPA', 'Marks', 'Achievements', 'Other Info'
    ]
    for column in columns_to_add:
        if column not in df.columns:
            df[column] = ""
    return df

def write_log(message):
    with open(log_file_path, 'a') as log_file:
        log_file.write(f"{datetime.now()} - {message}\n")

def log_processing_status(success_count, failure_count):
    success_rate = (success_count / (success_count + failure_count)) * 100 if (success_count + failure_count) > 0 else 0
    failure_rate = 100 - success_rate
    log_message = f"Resumes processed: Success Rate: {success_rate:.2f}%, Failure Rate: {failure_rate:.2f}%"
    write_log(log_message)

def parse_resume(resume_text):
    print("Parsing resume text...")
    extracted_data = {
        **extract_section(resume_text, ["Experience", "Work Experience"]),
        **extract_section(resume_text, ["Skills", "Technical Skills", "Key Skills"]),
        **extract_section(resume_text, ["Projects", "Key Projects"]),
        **extract_section(resume_text, ["Department", "Field"]),
        **extract_section(resume_text, [r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", "Contact Email"])
        **extract_section(resume_text, [r"https?://(?:www\.)?linkedin\.com/in/[a-zA-Z0-9-]+", "LinkedIn Profile"])
        **extract_section(resume_text, ["CGPA", "Cumulative GPA"]),
        **extract_section(resume_text, ["Marks", "Academic Performance"]),
        **extract_section(resume_text, ["Achievements", "Key Achievements"]),
        **extract_section(resume_text, ["Other Info", "Additional Information"]),
    }
    fallback_data = []
    required_keys = [
        "Experience", "Skills", "Projects", "Department",
        "Email", "LinkedIn", "CGPA", "Marks", "Achievements", "Other Info"
    ]
    
    for key in required_keys:
        if key not in extracted_data:
            extracted_data[key] = "Not Found"
    
    extracted_data["Fallback"] = "; ".join(fallback_data) if fallback_data else "No extra data"
    print("Extracted Data:", extracted_data)
    return extracted_data

state_file = 'parsed_resumes.json'

import json
import pandas as pd

def load_parsed_state():
    try:
        with open(state_file, 'r') as file:
            return json.load(file)
    except FileNotFoundError:
        return {}

def save_parsed_state(parsed_state):
    with open(state_file, 'w') as file:
        json.dump(parsed_state, file)

parsed_state = load_parsed_state()

@app.route('/process', methods=['POST'])
def process_resumes(employee_id_filter=None):
    try:
        write_log("Starting resume processing...")
        write_log("Authenticating Google Sheets...")
        service = authenticate_google_sheets()

        range_ = f"{SHEET_NAME}!A1:Z1000"
        write_log(f"Fetching data from range: {range_}")
        result = service.spreadsheets().values().get(spreadsheetId=SPREADSHEET_ID, range=range_).execute()
        values = result.get('values', [])

        if not values:
            write_log("No data found in the sheet.")
            return jsonify({"message": "No data found in the sheet."}), 400

        headers = values[0]
        write_log(f"Sheet Headers: {headers}")
        if "Resume" not in headers or "Employee ID" not in headers:
            write_log("Expected columns not found in sheet.")
            return jsonify({"message": "Expected columns not found: 'Resume' and 'Employee ID'."}), 400

        data = values[1:]
        df = pd.DataFrame(data, columns=headers)
        write_log("Initial DataFrame:")
        print(df.head())

        if employee_id_filter:
            df = df[df["Employee ID"].isin(employee_id_filter) if isinstance(employee_id_filter, list) else df["Employee ID"] == employee_id_filter]
            write_log(f"Filtered DataFrame based on Employee IDs: {employee_id_filter}")
            print(df.head())

        if df.empty:
            write_log("No matching Employee ID(s) found.")
            return jsonify({"message": "No matching Employee ID(s) found."}), 404

        success_count = 0
        failure_count = 0
        parsed_resumes = []

        batch_size = 50
        for i in range(0, len(df), batch_size):
            batch_df = df.iloc[i:i + batch_size]

            for _, row in batch_df.iterrows():
                employee_id = row.get("Employee ID", "Unknown")
                resume_link = row.get("Resume", "")
                write_log(f"Processing resume for Employee ID: {employee_id}, Resume: {resume_link}")

                if resume_link and "http" in str(resume_link):
                    pdf_path = download_pdf_from_drive(resume_link)
                    write_log(f"Downloaded PDF Path: {pdf_path}")
                    
                    if parsed_state.get(pdf_path, False):
                        write_log(f"Skipping {pdf_path}, already parsed.")
                        continue 
                    
                    if pdf_path:
                        resume_text = extract_text_from_pdf(pdf_path)
                        write_log(f"Extracted text for Employee ID {employee_id}: {resume_text[:100]}...")

                        parsed_data = extract_section(resume_text, ["Experience", "Skills", "Projects", "Department", "Email", "LinkedIn", "CGPA", "Marks", "Achievements", "Other Info"])
                        
                        experience = str(parsed_data.get('Experience', 'Not found'))
                        skills = str(parsed_data.get('Skills', 'Not found'))
                        department = str(parsed_data.get('Department', 'Not found'))
                        projects = str(parsed_data.get('Projects', 'Not found'))
                        email = str(parsed_data.get('Email', 'Not found'))
                        linkedin = str(parsed_data.get('LinkedIn', 'Not found'))
                        cgpa = str(parsed_data.get('CGPA', 'Not found'))
                        marks = str(parsed_data.get('Marks', 'Not found'))
                        achievements = str(parsed_data.get('Achievements', 'Not found'))
                        other_info = str(parsed_data.get('Other Info', 'Not found'))

                        parsed_resumes.append({
                            "Employee ID": employee_id,
                            "Experience": experience,
                            "Skills": skills,
                            "Department": department,
                            "Projects": projects,
                            "Email": email,
                            "LinkedIn": linkedin,
                            "CGPA": cgpa,
                            "Marks": marks,
                            "Achievements": achievements,
                            "Other Info": other_info
                        })
                        success_count += 1
                    else:
                        failure_count += 1
                        write_log(f"Failed to download or process resume for Employee ID: {employee_id}")
                else:
                    failure_count += 1
                    write_log(f"Invalid resume link for Employee ID: {employee_id}")

            if parsed_resumes:
                parsed_state[pdf_path] = True 
                save_parsed_state(parsed_state)

                parsed_df = pd.DataFrame(parsed_resumes)
                write_log("Parsed Resume DataFrame for batch:")
                print(parsed_df.head())
                
                update_google_sheet_with_parsed_data(service, SPREADSHEET_ID, SHEET_NAME, parsed_df)

            parsed_resumes = [] 

        write_log("Resumes processed successfully!")
        return jsonify({
            "message": "Resumes processed successfully!",
            "parsed_resumes": parsed_resumes
        }), 200

    except Exception as e:
        write_log(f"Error: {str(e)}")
        return jsonify({"error": str(e)}), 500

port= 9090
public_url = ngrok.connect(port)
print(f" * ngrok tunnel \"{public_url}\" -> \"http://127.0.0.1:{port}\"")
app.run(port=port)